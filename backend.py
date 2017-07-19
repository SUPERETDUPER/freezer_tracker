import datetime
import os.path

import openpyxl

import globalvar

db_file_path = "freezer_inventory_database.xlsx"

idColumn = "Batch number"
timeColumn = "Time stamp"
typeColumn = "Type"
subtypeColumn = "Sub-type"
weightColumn = "Weight"

columns = {idColumn: 0,
           timeColumn: 1,
           typeColumn: 2,
           subtypeColumn: 3,
           weightColumn: 4}


class Database:
    def __init__(self):
        # TODO : Implement edge case of reaching id 99 999

        if os.path.isfile(db_file_path):
            self.workbook = openpyxl.load_workbook(filename=db_file_path)  # If file exists, load it
            self.ws = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()  # Else create it and create headers
            self.ws = self.workbook.active

            self.create_header()

        self.last_id = 10000

        for index, row in enumerate(self.ws.iter_rows(row_offset=1)):  # Loop through every row

            empty = True

            for cell in row:
                if cell.value != "" and cell.value is not None:
                    empty = False

            if empty:  # If empty tracker is still True after looping through all the row's cells, this line is empty
                self.lastRow = index
                break

            if self.last_id < row[columns[idColumn]].value:  # Update to get latest id assigned
                self.last_id = row[columns[idColumn]].value

    def create_header(self):
        for column in columns.keys():
            header_cell = self.ws.cell(row=1, column=columns[column] + 1)
            header_cell.value = column

    def save(self):
        self.workbook.save(filename=db_file_path)

    def get_info(self):
        print("Last row : " + str(self.lastRow))
        print("Last id :" + str(self.last_id))

    def add_item(self, row):
        return 0

    def remove_item(self, batch_number):
        return True

    def get_row(self, batch_number):
        return Row()


class Row:
    def __init__(self, category=None, subcategory=None, weight=None, batch_number=None, entry_date=None):
        if entry_date is None:
            entry_date = '{:%Y-%m-%d %H:%M}'.format(datetime.datetime.now())

        self.row = [None] * len(columns)

        self.row[columns[idColumn]] = batch_number
        self.row[columns[timeColumn]] = entry_date
        self.row[columns[typeColumn]] = category
        self.row[columns[subtypeColumn]] = subcategory
        self.row[columns[weightColumn]] = weight

    def iter_row(self):
        return self.row

    def get_item(self, index):
        return self.row[index]


def get_letter(index):
    return chr(ord("A") + index)


globalvar.database = Database()
globalvar.database.save()
globalvar.database.get_info()
